from __future__ import annotations

import json
import re
from collections import OrderedDict, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
COMPARISON_FILE = ROOT / "medicine_comparison_22-07-2026.xlsx"
GENERATED_MASTER_FILE = ROOT / "backend" / "src" / "data" / "generated" / "godownInventory.generated.js"
JULY_MIGRATION_FILE = ROOT / "backend" / "src" / "database" / "migrations" / "022_godown_medicines_22_07_26.sql"
OUTPUT_DIR = ROOT / "deliverables"
OUTPUT_FILE = OUTPUT_DIR / "HMS_Consolidated_Medicine_Master_2026-08-20.xlsx"


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def normalized(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())


def unique_join(values, separator=" | ") -> str:
    result = []
    seen = set()
    for value in values:
        text = clean(value)
        key = text.casefold()
        if text and key not in seen:
            seen.add(key)
            result.append(text)
    return separator.join(result)


def load_generated_masters() -> list[dict]:
    source = GENERATED_MASTER_FILE.read_text(encoding="utf-8")
    payload = source[source.index("=") + 1 :].strip().rstrip(";")
    data = json.loads(payload)
    records = []
    for item in data["medicineMasters"]:
        records.append(
            {
                "code": clean(item.get("code")),
                "name": clean(item.get("name")),
                "company": clean(item.get("category")),
                "formulation": clean(item.get("formulation")),
                "unit": clean(item.get("unit")) or "unit",
                "packaging": clean(item.get("defaultPackaging")),
                "purchase_price": item.get("purchasePrice", 0) or 0,
                "selling_price": item.get("price", item.get("sellingPrice", 0)) or 0,
                "reorder_level": item.get("reorderLevel", 0) or 0,
                "gst": item.get("gstPercent", 0) or 0,
                "hsn": clean(item.get("hsnCode")),
                "origin": "March godown master",
            }
        )
    return records


def load_migration_masters() -> list[dict]:
    source = JULY_MIGRATION_FILE.read_text(encoding="utf-8")
    row_pattern = re.compile(
        r"\('[^']+'::uuid, '([^']+)', '((?:''|[^'])*)', '((?:''|[^'])*)', "
        r"'((?:''|[^'])*)', '((?:''|[^'])*)', '(\{.*?\})'::jsonb\)"
    )
    records = []
    for code, name, company, formulation, unit, metadata_text in row_pattern.findall(source):
        metadata = json.loads(metadata_text)
        records.append(
            {
                "code": clean(code),
                "name": clean(name.replace("''", "'")),
                "company": clean(company.replace("''", "'")),
                "formulation": clean(formulation),
                "unit": clean(unit) or "unit",
                "packaging": clean(metadata.get("packing")),
                "purchase_price": 0,
                "selling_price": 0,
                "reorder_level": 0,
                "gst": 0,
                "hsn": "",
                "origin": "July medicine migration",
            }
        )
    return records


def hospital_masters() -> list[dict]:
    # These five non-GDN medicine masters are seeded separately in store.js.
    return [
        {
            "code": "SRA-MED-001",
            "name": "Mahayograj Guggulu",
            "company": "Ayurvedic Classical",
            "formulation": "Tablet",
            "unit": "tablet",
            "selling_price": 18,
            "reorder_level": 40,
            "gst": 5,
        },
        {
            "code": "SRA-MED-002",
            "name": "Dashmool Kwath",
            "company": "Ayurvedic Classical",
            "formulation": "Kwath",
            "unit": "bottle",
            "selling_price": 140,
            "reorder_level": 20,
            "gst": 5,
        },
        {
            "code": "SRA-MED-003",
            "name": "Ashwagandha Churna",
            "company": "Ayurvedic Classical",
            "formulation": "Churna",
            "unit": "jar",
            "selling_price": 165,
            "reorder_level": 15,
            "gst": 5,
        },
        {
            "code": "SRA-MED-004",
            "name": "Brahmi Vati",
            "company": "Ayurvedic Classical",
            "formulation": "Tablet",
            "unit": "tablet",
            "selling_price": 14,
            "reorder_level": 35,
            "gst": 5,
        },
        {
            "code": "SRA-MED-005",
            "name": "Nirgundi Taila",
            "company": "External Therapy",
            "formulation": "Taila",
            "unit": "bottle",
            "selling_price": 220,
            "reorder_level": 12,
            "gst": 12,
        },
    ]


def load_latest_catalog(workbook) -> OrderedDict[tuple[str, str], dict]:
    catalog: OrderedDict[tuple[str, str], dict] = OrderedDict()
    sheet = workbook["All Latest Rows"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        (
            source_sheet,
            source_row,
            company,
            name,
            botanical,
            packaging,
            quantity,
            location,
            formulation,
            _comparison_status,
            quantity_issue,
            extraction_notes,
        ) = row[:12]
        if not clean(company) or not clean(name):
            continue
        key = (normalized(company), normalized(name))
        if key not in catalog:
            catalog[key] = {
                "name": clean(name),
                "company": clean(company),
                "formulation": clean(formulation) or "General",
                "botanical_values": [],
                "packaging_values": [],
                "quantity_values": [],
                "location_values": [],
                "source_values": [],
                "issue_values": [],
            }
        item = catalog[key]
        item["botanical_values"].append(botanical)
        item["packaging_values"].append(packaging)
        item["quantity_values"].append(quantity)
        item["location_values"].append(location)
        item["source_values"].append(f"{clean(source_sheet)}!{source_row}")
        item["issue_values"].extend([quantity_issue, extraction_notes])
    return catalog


def load_review_mapping(workbook) -> tuple[dict[tuple[str, str], str], set[str]]:
    latest_key_to_db_code = {}
    superseded_codes = set()
    sheet = workbook["Review Matches"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        db_code, latest_company, latest_name = clean(row[3]), clean(row[6]), clean(row[7])
        if db_code:
            superseded_codes.add(db_code)
        if db_code and latest_company and latest_name:
            latest_key_to_db_code[(normalized(latest_company), normalized(latest_name))] = db_code
    return latest_key_to_db_code, superseded_codes


def load_valid_legacy_codes(workbook, superseded_codes: set[str]) -> list[str]:
    codes = []
    for row in workbook["DB Only"].iter_rows(min_row=2, values_only=True):
        code = clean(row[0])
        if code and code not in superseded_codes:
            codes.append(code)
    return codes


def infer_unit(packaging: str) -> str:
    text = clean(packaging).upper()
    units = (
        ("ML", "ml"),
        ("LTR", "ltr"),
        ("LITRE", "ltr"),
        ("KG", "kg"),
        ("GM", "gm"),
        ("MG", "mg"),
        ("TAB", "tab"),
        ("CAP", "cap"),
        ("NOS", "nos"),
        ("BOX", "box"),
        ("PACK", "pack"),
        ("BOTTLE", "bottle"),
        ("JAR", "jar"),
    )
    for token, unit in units:
        if token in text:
            return unit
    return "unit"


def optional_number(value):
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        return None
    return number if number != 0 else None


def build_consolidated_records() -> tuple[list[dict], dict]:
    comparison = load_workbook(COMPARISON_FILE, data_only=True, read_only=True)
    current_godown = load_generated_masters() + load_migration_masters()
    current_by_code = {item["code"]: item for item in current_godown}
    current_by_key = defaultdict(list)
    current_by_name = defaultdict(list)
    for item in current_godown:
        current_by_key[(normalized(item["company"]), normalized(item["name"]))].append(item)
        current_by_name[normalized(item["name"])].append(item)

    latest_catalog = load_latest_catalog(comparison)
    review_mapping, superseded_codes = load_review_mapping(comparison)
    legacy_codes = load_valid_legacy_codes(comparison, superseded_codes)

    records = []
    used_codes = set()
    counts = defaultdict(int)

    # Resolve current codes in stable passes. Exact company/name matches have
    # priority, then curated review corrections, then unique cross-category
    # matches. This prevents one current code from being assigned to two rows
    # when the latest workbook legitimately lists the same name for two firms.
    assignments = {}
    assignment_types = {}
    for key in latest_catalog:
        matches = current_by_key.get(key, [])
        if len(matches) == 1:
            assignments[key] = matches[0]
            assignment_types[key] = "exact"
            used_codes.add(matches[0]["code"])

    for key in latest_catalog:
        if key in assignments:
            continue
        review_code = review_mapping.get(key)
        if review_code and review_code in current_by_code and review_code not in used_codes:
            assignments[key] = current_by_code[review_code]
            assignment_types[key] = "corrected"
            used_codes.add(review_code)

    for key, latest in latest_catalog.items():
        if key in assignments:
            continue
        available_same_name = [
            item for item in current_by_name.get(normalized(latest["name"]), []) if item["code"] not in used_codes
        ]
        if len(available_same_name) == 1:
            assignments[key] = available_same_name[0]
            assignment_types[key] = "normalized"
            used_codes.add(available_same_name[0]["code"])

    for key, latest in latest_catalog.items():
        match = assignments.get(key)
        match_type = assignment_types.get(key, "new")

        packaging = unique_join(latest["packaging_values"])
        note_parts = []
        if match_type in {"corrected", "normalized"} and match:
            old_label = f'{match["company"]} / {match["name"]}'
            new_label = f'{latest["company"]} / {latest["name"]}'
            if normalized(old_label) != normalized(new_label):
                note_parts.append(f"HMS label normalized from: {old_label}")
        issues = unique_join(latest["issue_values"], "; ")
        if issues:
            note_parts.append(issues)

        if match:
            counts["current_latest"] += 1
            status = "CURRENT HMS"
            if match_type == "corrected":
                status = "CURRENT HMS - CORRECTED"
            elif match_type == "normalized":
                status = "CURRENT HMS - NORMALIZED"
        else:
            counts["new"] += 1
            status = "NEW - ADD TO HMS"

        records.append(
            {
                "code": match["code"] if match else "",
                "status": status,
                "name": latest["name"],
                "company": latest["company"],
                "formulation": latest["formulation"] or (match["formulation"] if match else "General"),
                "unit": (match["unit"] if match else infer_unit(packaging)) or "unit",
                "packaging": packaging or (match["packaging"] if match else ""),
                "purchase_price": optional_number(match.get("purchase_price")) if match else None,
                "selling_price": optional_number(match.get("selling_price")) if match else None,
                "stock_quantity": None,
                "reorder_level": optional_number(match.get("reorder_level")) if match else None,
                "gst": optional_number(match.get("gst")) if match else None,
                "hsn": match.get("hsn", "") if match else "",
                "reference_stock": unique_join(latest["quantity_values"]),
                "location": unique_join(latest["location_values"]),
                "botanical": unique_join(latest["botanical_values"]),
                "source_notes": "; ".join(note_parts),
                "source_rows": unique_join(latest["source_values"], "; "),
            }
        )

    for code in legacy_codes:
        item = current_by_code[code]
        if code in used_codes:
            raise ValueError(f"Legacy code was already assigned to a latest record: {code}")
        used_codes.add(code)
        counts["legacy"] += 1
        records.append(
            {
                "code": code,
                "status": "CURRENT HMS - LEGACY",
                "name": item["name"],
                "company": item["company"],
                "formulation": item["formulation"],
                "unit": item["unit"],
                "packaging": item["packaging"],
                "purchase_price": optional_number(item.get("purchase_price")),
                "selling_price": optional_number(item.get("selling_price")),
                "stock_quantity": None,
                "reorder_level": optional_number(item.get("reorder_level")),
                "gst": optional_number(item.get("gst")),
                "hsn": item.get("hsn", ""),
                "reference_stock": "",
                "location": "",
                "botanical": "",
                "source_notes": "Valid older HMS medicine not present in the July workbook; retained.",
                "source_rows": "March HMS medicine master",
            }
        )

    for item in hospital_masters():
        counts["hospital"] += 1
        records.append(
            {
                "code": item["code"],
                "status": "CURRENT HMS - HOSPITAL",
                "name": item["name"],
                "company": item["company"],
                "formulation": item["formulation"],
                "unit": item["unit"],
                "packaging": "",
                "purchase_price": None,
                "selling_price": optional_number(item["selling_price"]),
                "stock_quantity": None,
                "reorder_level": optional_number(item["reorder_level"]),
                "gst": optional_number(item["gst"]),
                "hsn": "",
                "reference_stock": "",
                "location": "",
                "botanical": "",
                "source_notes": "Separate hospital medicine master from store.js.",
                "source_rows": "HMS hospital seed master",
            }
        )

    records.sort(key=lambda item: (item["company"].casefold(), item["name"].casefold(), item["code"]))

    final_keys = [(normalized(item["company"]), normalized(item["name"])) for item in records]
    if len(final_keys) != len(set(final_keys)):
        duplicates = [key for key in set(final_keys) if final_keys.count(key) > 1]
        raise ValueError(f"Final catalog contains duplicate company/medicine keys: {duplicates[:5]}")
    if len(records) != 1169:
        raise ValueError(f"Expected 1,169 consolidated medicines, found {len(records)}")
    if counts != {"current_latest": 1017, "new": 92, "legacy": 55, "hospital": 5}:
        raise ValueError(f"Unexpected catalog breakdown: {dict(counts)}")
    if len(used_codes) != 1072:
        raise ValueError(f"Expected 1,072 valid current GDN codes, found {len(used_codes)}")

    summary = {
        "total": len(records),
        "current_godown": len(used_codes),
        "current_hospital": counts["hospital"],
        "new": counts["new"],
        "latest_unique": len(latest_catalog),
        "legacy": counts["legacy"],
        "superseded_removed": len(current_godown) - len(used_codes),
    }
    return records, summary


def make_workbook(records: list[dict], summary: dict) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Medicine Master"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "D5"
    sheet.sheet_properties.tabColor = "1F4E78"

    navy = "1F4E78"
    blue = "5B9BD5"
    pale_blue = "DDEBF7"
    pale_yellow = "FFF2CC"
    pale_green = "E2F0D9"
    pale_orange = "FCE4D6"
    pale_gray = "E7E6E6"
    white = "FFFFFF"
    thin_gray = Side(style="thin", color="B7B7B7")

    sheet.merge_cells("A1:S1")
    sheet["A1"] = "HMS Consolidated Medicine Master - Current, Corrected and New"
    sheet["A1"].font = Font(name="Calibri", size=16, bold=True, color=white)
    sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells("A2:S2")
    sheet["A2"] = (
        f"{summary['total']:,} unique medicines: {summary['current_godown'] + summary['current_hospital']:,} valid current HMS records "
        f"and {summary['new']:,} new records. Enter/update the yellow price, stock, reorder, GST and HSN cells."
    )
    sheet["A2"].font = Font(name="Calibri", size=10, italic=True, color="404040")
    sheet["A2"].fill = PatternFill("solid", fgColor=pale_blue)
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[2].height = 22

    sheet.merge_cells("A3:S3")
    sheet["A3"] = (
        "Reference stock/location comes from the 22-Jul-2026 workbook and may contain mixed units; use Stock Quantity for the verified current figure. "
        "Seven invalid serial-number medicines were removed."
    )
    sheet["A3"].font = Font(name="Calibri", size=9, color="7F6000")
    sheet["A3"].fill = PatternFill("solid", fgColor=pale_yellow)
    sheet["A3"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[3].height = 22

    headers = [
        "S.No.",
        "HMS Medicine Code",
        "Record Status",
        "Medicine Name",
        "Company / Category",
        "Formulation",
        "Unit",
        "Packaging",
        "Purchase Price (INR)",
        "Selling Price (INR)",
        "Stock Quantity",
        "Reorder Level",
        "GST %",
        "HSN Code",
        "Last Recorded Stock (22-Jul-2026)",
        "Last Recorded Location",
        "Botanical Name",
        "Notes",
        "Source Rows",
    ]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=column, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy if column < 9 or column > 14 else blue)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    sheet.row_dimensions[4].height = 38

    comments = {
        2: "Blank for new medicines that do not yet have a medicine_code in the HMS codebase.",
        9: "Editable. Enter the purchase rate per listed unit.",
        10: "Editable. Enter the selling/dispensing rate per listed unit.",
        11: "Editable. Enter verified current stock in the listed unit.",
        15: "Reference text copied from the July workbook. Mixed expressions were deliberately not converted into a potentially incorrect number.",
    }
    for column, comment in comments.items():
        sheet.cell(4, column).comment = Comment(comment, "OpenAI Codex")

    for row_number, record in enumerate(records, 5):
        values = [
            row_number - 4,
            record["code"],
            record["status"],
            record["name"],
            record["company"],
            record["formulation"],
            record["unit"],
            record["packaging"],
            record["purchase_price"],
            record["selling_price"],
            record["stock_quantity"],
            record["reorder_level"],
            record["gst"],
            record["hsn"],
            record["reference_stock"],
            record["location"],
            record["botanical"],
            record["source_notes"],
            record["source_rows"],
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=column >= 15)
            cell.border = Border(bottom=Side(style="hair", color="D9E1F2"))
            if 9 <= column <= 14:
                cell.fill = PatternFill("solid", fgColor=pale_yellow)
        sheet.cell(row_number, 1).alignment = Alignment(horizontal="center", vertical="top")
        sheet.cell(row_number, 3).alignment = Alignment(horizontal="center", vertical="top")
        for column in (9, 10):
            sheet.cell(row_number, column).number_format = '#,##0.00'
        for column in (11, 12):
            sheet.cell(row_number, column).number_format = '#,##0.00'
        sheet.cell(row_number, 13).number_format = '0.00'

    last_row = len(records) + 4
    table = Table(displayName="MedicineMaster", ref=f"A4:S{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    sheet.add_table(table)

    nonnegative_decimal = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    nonnegative_decimal.error = "Enter a number greater than or equal to 0."
    nonnegative_decimal.errorTitle = "Invalid number"
    nonnegative_decimal.prompt = "Enter a verified non-negative value, or leave blank."
    nonnegative_decimal.promptTitle = "Editable field"
    nonnegative_decimal.showErrorMessage = True
    nonnegative_decimal.showInputMessage = True
    sheet.add_data_validation(nonnegative_decimal)
    nonnegative_decimal.add(f"I5:M{last_row}")

    sheet.conditional_formatting.add(
        f"C5:C{last_row}",
        FormulaRule(formula=['LEFT($C5,3)="NEW"'], fill=PatternFill("solid", fgColor=pale_orange)),
    )
    sheet.conditional_formatting.add(
        f"C5:C{last_row}",
        FormulaRule(formula=['LEFT($C5,7)="CURRENT"'], fill=PatternFill("solid", fgColor=pale_green)),
    )
    sheet.conditional_formatting.add(
        f"B5:B{last_row}",
        FormulaRule(formula=['$B5=""'], fill=PatternFill("solid", fgColor=pale_gray)),
    )

    widths = {
        "A": 8,
        "B": 19,
        "C": 25,
        "D": 34,
        "E": 24,
        "F": 16,
        "G": 11,
        "H": 20,
        "I": 18,
        "J": 18,
        "K": 16,
        "L": 15,
        "M": 10,
        "N": 14,
        "O": 29,
        "P": 25,
        "Q": 30,
        "R": 45,
        "S": 34,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.auto_filter.ref = f"A4:S{last_row}"
    sheet.sheet_view.zoomScale = 85
    sheet.print_title_rows = "1:4"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "Page &P of &N"
    sheet.oddFooter.right.text = "Generated 20-Aug-2026"

    workbook.properties.title = "HMS Consolidated Medicine Master"
    workbook.properties.subject = "Single-sheet current, corrected, legacy, hospital and new medicine catalog"
    workbook.properties.creator = "OpenAI Codex"
    workbook.properties.created = datetime(2026, 8, 20)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_FILE)


def validate_output(expected_records: list[dict]) -> dict:
    workbook = load_workbook(OUTPUT_FILE, data_only=False, read_only=False)
    if workbook.sheetnames != ["Medicine Master"]:
        raise ValueError(f"Workbook must contain exactly one sheet; found {workbook.sheetnames}")
    sheet = workbook["Medicine Master"]
    data_rows = sheet.max_row - 4
    if data_rows != len(expected_records):
        raise ValueError(f"Expected {len(expected_records)} data rows, found {data_rows}")
    codes = [clean(sheet.cell(row, 2).value) for row in range(5, sheet.max_row + 1)]
    statuses = [clean(sheet.cell(row, 3).value) for row in range(5, sheet.max_row + 1)]
    names = [clean(sheet.cell(row, 4).value) for row in range(5, sheet.max_row + 1)]
    companies = [clean(sheet.cell(row, 5).value) for row in range(5, sheet.max_row + 1)]
    keys = [(normalized(company), normalized(name)) for company, name in zip(companies, names)]
    nonblank_codes = [code for code in codes if code]
    if len(nonblank_codes) != len(set(nonblank_codes)):
        raise ValueError("Workbook contains duplicate nonblank HMS medicine codes")
    if len(keys) != len(set(keys)):
        raise ValueError("Workbook contains duplicate normalized company/medicine rows")
    if any(not name or not company for name, company in zip(names, companies)):
        raise ValueError("Workbook contains a blank medicine or company")
    return {
        "file": str(OUTPUT_FILE),
        "sheets": len(workbook.sheetnames),
        "rows": data_rows,
        "coded_current_rows": len(nonblank_codes),
        "new_rows": sum(status.startswith("NEW") for status in statuses),
        "size_bytes": OUTPUT_FILE.stat().st_size,
    }


def main() -> None:
    records, summary = build_consolidated_records()
    make_workbook(records, summary)
    validation = validate_output(records)
    print(json.dumps({"summary": summary, "validation": validation}, indent=2))


if __name__ == "__main__":
    main()

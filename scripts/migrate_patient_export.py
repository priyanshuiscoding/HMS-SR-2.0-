#!/usr/bin/env python3
"""
Import and reconcile the cleaned HMS patient Excel export.

Default mode is a dry run. Pass --apply to write updates/inserts.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import uuid
from datetime import date, datetime
from pathlib import Path

import psycopg2
from openpyxl import load_workbook
from psycopg2.extras import Json, RealDictCursor


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXPORT = Path(r"C:\Users\HP\Downloads\patient_export_cleaned.xlsx")
DEFAULT_REGISTRATION_DATE = date(2026, 1, 1)
TITLES = {
    "baby": "Baby",
    "br": "Br",
    "dr": "Dr",
    "ku": "Ku",
    "master": "Master",
    "miss": "Miss",
    "mr": "Mr",
    "mrs": "Mrs",
    "ms": "Ms",
}
EMPTY_VALUES = {"", "0", "00", "000", "na", "n/a", "nil", "none", "null", "undefined", "-"}


def load_env() -> None:
    for env_file in (ROOT / "backend/.env", ROOT / ".env", ROOT / ".env.production"):
        if not env_file.exists():
            continue
        for line in env_file.read_text(encoding="utf-8").splitlines():
            text = line.strip()
            if not text or text.startswith("#") or "=" not in text:
                continue
            key, value = text.split("=", 1)
            os.environ.setdefault(key.strip(), value.strip())


def clean_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    text = str(value).replace("`", "").strip()
    return "" if text.lower() in {"nil", "none", "null", "undefined"} else text


def clean_nullable(value) -> str:
    text = clean_cell(value)
    return "" if text.lower() in EMPTY_VALUES else text


def clean_phone(value) -> str:
    return re.sub(r"[^\d+]", "", clean_nullable(value))


def normalize_gender(value) -> str:
    text = clean_nullable(value).lower()
    if text in {"m", "male"}:
        return "male"
    if text in {"f", "female"}:
        return "female"
    return "other" if text else ""


def parse_datetime(value):
    if isinstance(value, datetime):
        return value
    text = clean_nullable(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def parse_date(value):
    parsed = parse_datetime(value)
    return parsed.date() if parsed else None


def parse_age_days(value):
    text = clean_nullable(value)
    if not text:
        return None
    try:
        days = int(float(text))
    except ValueError:
        return None
    if days <= 0:
        return None
    return max(0, int(days / 365.2425))


def stable_uuid(source: str) -> str:
    digest = hashlib.sha1(source.encode("utf-8")).hexdigest()[:32]
    return str(uuid.UUID(digest))


def normalize_title(value) -> str:
    text = clean_nullable(value).lower().replace(".", "")
    return TITLES.get(text, clean_nullable(value).title())


def split_name(row: dict) -> dict:
    title = normalize_title(row.get("salutation")) or normalize_title(row.get("title"))
    full_name = re.sub(r"\s+", " ", clean_nullable(row.get("full_name"))).strip()
    first_name = re.sub(r"\s+", " ", clean_nullable(row.get("first_name"))).strip()
    middle_name = re.sub(r"\s+", " ", clean_nullable(row.get("middle_name"))).strip()
    last_name = re.sub(r"\s+", " ", clean_nullable(row.get("last_name"))).strip()

    if not full_name:
        full_name = " ".join(part for part in (title, first_name, middle_name, last_name) if part)

    without_title = full_name
    if title and without_title.lower().startswith(f"{title.lower()} "):
        without_title = without_title[len(title):].strip()

    if not first_name:
        first_name = without_title.split(" ", 1)[0] if without_title else "Unknown"

    if not last_name:
        first_parts = first_name.split()
        remaining = without_title.split()
        if remaining[: len(first_parts)] == first_parts:
            last_name = " ".join(remaining[len(first_parts):])

    return {
        "title": title,
        "first_name": first_name or "Unknown",
        "last_name": last_name,
        "full_name": full_name or "Unknown",
    }


def parse_address(row: dict) -> dict:
    parsed = {}
    raw_address = clean_cell(row.get("address"))
    if raw_address.startswith("{") and raw_address.endswith("}"):
        try:
            parsed = json.loads(raw_address)
        except json.JSONDecodeError:
            parsed = {}

    house = clean_nullable(row.get("address_line1")) or clean_nullable(parsed.get("addressLine1"))
    line2 = clean_nullable(row.get("address_line2")) or clean_nullable(parsed.get("addressLine2"))
    line3 = clean_nullable(row.get("address_line3")) or clean_nullable(parsed.get("addressLine3"))
    city = clean_nullable(row.get("address_city")) or clean_nullable(parsed.get("city")) or clean_nullable(row.get("city"))
    state = clean_nullable(row.get("address_state")) or clean_nullable(parsed.get("state")) or "Madhya Pradesh"
    pincode = clean_nullable(row.get("address_pinCode")) or clean_nullable(parsed.get("pinCode"))
    area = ", ".join(part for part in (line2, line3) if part)
    address = ", ".join(part for part in (house, area, city, state, pincode) if part)

    return {
        "address": address,
        "house_street": house,
        "area_village": area,
        "city": city,
        "state": state,
        "pincode": pincode,
    }


def read_records(export_path: Path, sheet: str) -> list[dict]:
    workbook = load_workbook(export_path, read_only=True, data_only=True)
    worksheet = workbook[sheet]
    rows = worksheet.iter_rows(values_only=True)
    header = [clean_cell(cell) for cell in next(rows)]
    records = []
    for row_number, row in enumerate(rows, start=2):
        record = {header[index]: clean_cell(row[index]) if index < len(row) else "" for index in range(len(header))}
        if any(record.values()):
            record["_sourceLine"] = row_number
            records.append(record)
    return records


def to_patient(row: dict, source_document: str) -> dict:
    ppin = clean_nullable(row.get("ppin"))
    names = split_name(row)
    address = parse_address(row)
    dob = parse_date(row.get("dob"))
    last_updated = parse_datetime(row.get("last_time_updated"))
    raw = {key: value for key, value in row.items() if not key.startswith("_")}
    raw["_sourceLine"] = row["_sourceLine"]

    metadata = {
        "oldPatientImport": True,
        "patientExportMigration": True,
        "ppin": ppin,
        "sourceLine": row["_sourceLine"],
        "legacyUhid": f"OLD-PPIN-{ppin.zfill(5)}" if ppin else "",
        "sourceUuid": clean_nullable(row.get("uuid")),
        "lastTimeUpdated": clean_nullable(row.get("last_time_updated")),
        "raw": raw,
    }

    return {
        "id": stable_uuid(f"patient-export:{ppin or row['_sourceLine']}:{names['full_name']}"),
        "uhid": "",
        "registration_number": ppin,
        "opd_ipd_number": "",
        "patient_type": "old",
        "title": names["title"],
        "first_name": names["first_name"],
        "last_name": names["last_name"],
        "full_name": names["full_name"],
        "father_name": clean_nullable(row.get("fathers_name")),
        "date_of_birth": dob,
        "age_years": parse_age_days(row.get("current_age_in_days")),
        "gender": normalize_gender(row.get("gender")),
        "blood_group": "",
        "marital_status": "",
        "occupation": "",
        "phone": clean_phone(row.get("phone_number")),
        "alt_phone": clean_phone(row.get("phone_number2")),
        "email": clean_nullable(row.get("email")),
        "address": address["address"],
        "house_street": address["house_street"],
        "area_village": address["area_village"],
        "city": address["city"],
        "state": address["state"],
        "pincode": address["pincode"],
        "id_type": "",
        "emergency_contact_name": "",
        "emergency_contact_phone": "",
        "registration_date": DEFAULT_REGISTRATION_DATE,
        "registration_time": None,
        "referred_by": "Cleaned patient export migration",
        "photo_url": "",
        "source_document": source_document,
        "clinical_notes": [],
        "created_by": None,
        "metadata": metadata,
        "last_updated": last_updated,
    }


def connect():
    load_env()
    return psycopg2.connect(
        host=os.environ.get("DB_HOST", "localhost"),
        port=os.environ.get("DB_PORT", "5432"),
        dbname=os.environ.get("DB_NAME", "hms_db"),
        user=os.environ.get("DB_USER", "hms_user"),
        password=os.environ.get("DB_PASSWORD", "hms_password"),
    )


def fetch_existing(cursor) -> dict[str, dict]:
    cursor.execute(
        """
        SELECT *
        FROM patients
        WHERE deleted_at IS NULL
        """
    )
    return {clean_cell(row["registration_number"]): dict(row) for row in cursor.fetchall() if clean_cell(row["registration_number"])}


def next_uhid(cursor) -> str:
    cursor.execute(
        """
        SELECT COALESCE(MAX(SUBSTRING(uhid FROM 6 FOR 6)::int), 0) + 1 AS next_number
        FROM patients
        WHERE uhid ~ '^SRH26[0-9]{6}$'
        """
    )
    return f"SRH26{int(cursor.fetchone()['next_number']):06d}"


def update_patient(cursor, existing: dict, patient: dict) -> None:
    merged_metadata = dict(existing.get("metadata") or {})
    merged_metadata.update(patient["metadata"])
    if existing.get("metadata", {}).get("uhidRemap"):
        merged_metadata["uhidRemap"] = existing["metadata"]["uhidRemap"]
    if existing.get("metadata", {}).get("legacyUhid"):
        merged_metadata["legacyUhid"] = existing["metadata"]["legacyUhid"]

    cursor.execute(
        """
        UPDATE patients
        SET
          registration_number = %(registration_number)s,
          opd_ipd_number = %(opd_ipd_number)s,
          patient_type = %(patient_type)s,
          title = %(title)s,
          first_name = %(first_name)s,
          last_name = %(last_name)s,
          full_name = %(full_name)s,
          father_name = %(father_name)s,
          date_of_birth = %(date_of_birth)s,
          age_years = %(age_years)s,
          gender = %(gender)s,
          blood_group = %(blood_group)s,
          marital_status = %(marital_status)s,
          occupation = %(occupation)s,
          phone = %(phone)s,
          alt_phone = %(alt_phone)s,
          email = %(email)s,
          address = %(address)s,
          house_street = %(house_street)s,
          area_village = %(area_village)s,
          city = %(city)s,
          state = %(state)s,
          pincode = %(pincode)s,
          id_type = %(id_type)s,
          emergency_contact_name = %(emergency_contact_name)s,
          emergency_contact_phone = %(emergency_contact_phone)s,
          referred_by = %(referred_by)s,
          source_document = %(source_document)s,
          metadata = %(metadata)s,
          updated_at = NOW()
        WHERE id = %(id)s
        """,
        {**patient, "id": existing["id"], "metadata": Json(merged_metadata)},
    )


def insert_patient(cursor, patient: dict) -> None:
    patient = {**patient, "uhid": next_uhid(cursor)}
    cursor.execute(
        """
        INSERT INTO patients (
          id, uhid, registration_number, opd_ipd_number, patient_type, title, first_name, last_name, full_name, father_name,
          date_of_birth, age_years, gender, blood_group, marital_status, occupation, phone, alt_phone, email,
          address, house_street, area_village, city, state, pincode, id_type, emergency_contact_name,
          emergency_contact_phone, registration_date, registration_time, referred_by, photo_url, source_document,
          clinical_notes, created_by, metadata
        )
        VALUES (
          %(id)s, %(uhid)s, %(registration_number)s, %(opd_ipd_number)s, %(patient_type)s, %(title)s, %(first_name)s,
          %(last_name)s, %(full_name)s, %(father_name)s, %(date_of_birth)s, %(age_years)s, %(gender)s, %(blood_group)s,
          %(marital_status)s, %(occupation)s, %(phone)s, %(alt_phone)s, %(email)s, %(address)s, %(house_street)s,
          %(area_village)s, %(city)s, %(state)s, %(pincode)s, %(id_type)s, %(emergency_contact_name)s,
          %(emergency_contact_phone)s, %(registration_date)s, %(registration_time)s, %(referred_by)s, %(photo_url)s,
          %(source_document)s, %(clinical_notes)s, %(created_by)s, %(metadata)s
        )
        """,
        {**patient, "clinical_notes": Json(patient["clinical_notes"]), "metadata": Json(patient["metadata"])},
    )


def name_anomalies(records: list[dict]) -> dict:
    columns = ["full_name", "first_name", "middle_name", "last_name", "fathers_name", "mothers_name", "salutation"]
    summary = {}
    for column in columns:
        values = [clean_cell(row.get(column)) for row in records]
        zeros = [row for row in records if clean_cell(row.get(column)).lower() in {"0", "00", "000"}]
        blanks = [row for row in records if not clean_cell(row.get(column))]
        summary[column] = {
            "zero": len(zeros),
            "blank": len(blanks),
            "sample": [
                {
                    "row": row["_sourceLine"],
                    "ppin": row.get("ppin", ""),
                    "full_name": row.get("full_name", ""),
                    column: row.get(column, ""),
                }
                for row in zeros[:10]
            ],
        }
    return summary


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("export", nargs="?", type=Path, default=DEFAULT_EXPORT)
    parser.add_argument("--sheet", default="CSV import")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    records = read_records(args.export, args.sheet)
    source_document = str(args.export)
    patients = [to_patient(row, source_document) for row in records]
    duplicate_ppins = len(patients) - len({patient["registration_number"] for patient in patients})

    with connect() as connection:
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            existing_by_ppin = fetch_existing(cursor)
            missing = [patient for patient in patients if patient["registration_number"] not in existing_by_ppin]
            existing = [patient for patient in patients if patient["registration_number"] in existing_by_ppin]

            if args.apply:
                for patient in existing:
                    update_patient(cursor, existing_by_ppin[patient["registration_number"]], patient)
                for patient in missing:
                    insert_patient(cursor, patient)
            else:
                connection.rollback()

    report = {
        "mode": "apply" if args.apply else "dry-run",
        "excelPatients": len(patients),
        "uniquePpins": len({patient["registration_number"] for patient in patients}),
        "duplicatePpins": duplicate_ppins,
        "existingMatchedByPpin": len(existing),
        "missingBeforeMigration": len(missing),
        "updatedExisting": len(existing) if args.apply else 0,
        "insertedMissing": len(missing) if args.apply else 0,
        "nameFieldAnomalies": name_anomalies(records),
        "missingSample": [
            {
                "ppin": patient["registration_number"],
                "fullName": patient["full_name"],
                "phone": patient["phone"],
            }
            for patient in missing[:20]
        ],
    }
    print(json.dumps(report, indent=2, default=str))


if __name__ == "__main__":
    main()
